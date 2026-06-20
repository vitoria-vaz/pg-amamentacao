import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CAMINHO_ENTRADA = r'C:\Users\vitoria-vaz\estudos\UFU\projeto-graduacao\pg-amamentacao\dataset_amamentacao_renomeado.csv'

# 1. Carregar os dados prontos
df = pd.read_csv(CAMINHO_ENTRADA, encoding='utf-8')

# Criar o arquivo Excel
wb = openpyxl.Workbook()

# =========================================================================
# TAB 1: APRESENTAÇÃO E INSTRUÇÕES
# =========================================================================
ws_intro = wb.active
ws_intro.title = "Apresentação"
ws_intro.views.sheetView[0].showGridLines = True

# Paleta de Cores (Azul Petróleo / Slate Blue Muted)
color_header = "204E70"      # Azul escuro corporativo
color_accent = "E6EDF2"      # Azul bem claro para blocos de destaque
color_zebra = "F5F8FA"       # Zebra striping suave
color_border = "D9D9D9"      # Borda cinza clara

font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
font_section = Font(name="Calibri", size=13, bold=True, color="204E70")
font_bold = Font(name="Calibri", size=11, bold=True)
font_regular = Font(name="Calibri", size=11, bold=False)
font_italic = Font(name="Calibri", size=10, italic=True, color="595959")

thin_border = Border(
    left=Side(style='thin', color=color_border),
    right=Side(style='thin', color=color_border),
    top=Side(style='thin', color=color_border),
    bottom=Side(style='thin', color=color_border)
)

# Título Principal
ws_intro["A2"] = "Base de Dados Processada - Estudo de Aleitamento Materno (ENANI-2019)"
ws_intro["A2"].font = font_title

ws_intro["A5"] = "Este arquivo contém o conjunto de dados finais perfeitamente limpos, estruturados e discretizados, pronto para uso em análises estatísticas e modelagem preditiva."
ws_intro["A5"].font = font_regular

# Quadro de Metadados
ws_intro["A7"] = "RESUMO DO DATASET"
ws_intro["A7"].font = font_section

metadata = [
    ["Métrica", "Valor", "Descrição"],
    ["Total de Instâncias (Linhas)", len(df), "Número de binômios (mãe-bebê) incluídos após os filtros de domínio."],
    ["Total de Atributos (Colunas)", len(df.columns), "Variáveis socioeconômicas, demográficas, clínicas e comportamentais."],
    ["Valores Ausentes (Nulos)", "0 (Zero)", "Todos os dados em branco foram tratados via Imputação Preditiva Encadeada."],
    ["Fonte Primária dos Dados", "ENANI-2019", "Pesquisa Nacional de Alimentação e Nutrição Infantil."],
    ["Status da Base", "Homologada", "Pronta para treinamento de modelos de Inteligência Artificial."]
]

for r_idx, row in enumerate(metadata, start=8):
    for c_idx, val in enumerate(row, start=1):
        cell = ws_intro.cell(row=r_idx, column=c_idx, value=val)
        cell.border = thin_border
        if r_idx == 8:
            cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        else:
            if c_idx == 1:
                cell.font = font_bold
            else:
                cell.font = font_regular
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color=color_zebra, end_color=color_zebra, fill_type="solid")

# Tabela de Distribuição do Alvo para dar contexto rápido
ws_intro["A16"] = "DISTRIBUIÇÃO DA VARIÁVEL ALVO (DESFECHO)"
ws_intro["A16"].font = font_section

alvo_counts = df['alvo_sucesso_ame_6m'].value_counts()
ws_intro["A18"] = "Classe Alvo"
ws_intro["B18"] = "Frequência Absoluta"
ws_intro["C18"] = "Frequência Relativa (%)"

for cell_id in ["A18", "B18", "C18"]:
    ws_intro[cell_id].fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    ws_intro[cell_id].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws_intro[cell_id].border = thin_border

row_num = 19
for cat, val in alvo_counts.items():
    ws_intro.cell(row=row_num, column=1, value=cat).font = font_regular
    ws_intro.cell(row=row_num, column=2, value=val).font = font_regular
    ws_intro.cell(row=row_num, column=3, value=f"{(val/len(df))*100:.2f}%").font = font_regular
    ws_intro.cell(row=row_num, column=1).border = thin_border
    ws_intro.cell(row=row_num, column=2).border = thin_border
    ws_intro.cell(row=row_num, column=3).border = thin_border
    row_num += 1

# Organizar larguras da aba intro
ws_intro.column_dimensions['A'].width = 32
ws_intro.column_dimensions['B'].width = 25
ws_intro.column_dimensions['C'].width = 50

# =========================================================================
# TAB 2: DICIONÁRIO DE VARIÁVEIS
# =========================================================================
ws_dic = wb.create_sheet(title="Dicionário de Variáveis")
ws_dic.views.sheetView[0].showGridLines = True

# Dados com a nova coluna "Observações Relevantes" detalhando o pré-processamento
dic_data = [
    ["Nome antigo (se tiver)","Nome do Atributo", "Descrição", "Categorias Disponíveis no Modelo", "Observações Relevantes (Pré-Processamento)"],
    ["a00_regiao", "regiao_residencia", "Região macroeconômica de residência da família", "Norte, Nordeste, Sudeste, Sul, Centro-Oeste", "Mantido o formato original do questionário."],
    ["a11_situacao", "zona_residencial", "Situação censitária do domicílio (Urbana ou Rural)", "Urbano, Rural", "Mantido o formato original do questionário."],
    ["alvo_amamentacao", "alvo_sucesso_ame_6m", "VARIÁVEL ALVO: Classificação binária do desfecho do Aleitamento Materno Exclusivo aos 6 meses", "Sucesso (AME 6m+), Desmame precoce (< 6m)", "Atributo derivado (Feature Engineering). Construído através do cruzamento de variáveis de tempo e alimentação."],
    ["escolaridade_cat", "escolaridade_mae", "Nível de instrução formal estruturado e agrupado da mãe", "Fundamental I, Fundamental II, Ensino Médio, Superior", "Agrupamento de categorias (Redução de Dimensionalidade). Ex: 'Sem estudo' fundido com 'Fundamental'."],
    ["filhos_vivos_cat", "qtd_filhos_vivos", "Histórico do número total de filhos nascidos vivos (Categorizado)", "Primípara (1), Multípara (2 ou mais)", "Discretizado e Imputado. Transformado em faixas binárias e valores nulos preenchidos com a Moda."],
    ["gestacoes_cat", "qtd_gestacoes", "Histórico do número total de gestações da mãe (Categorizado)", "Primípara (1), Multípara (2 ou mais)", "Discretizado e Imputado. Transformado em faixas binárias e valores nulos preenchidos com a Moda."],
    ["h04_parto", "tipo_parto", "Via ou modalidade cirúrgica/médica de realização do parto", "Normal, Cesariana agendada (eletiva), Cesariana de urgência (Não agendada)", "Mantido o formato original do questionário."],
    ["h05_chupeta_usou", "historico_uso_chupeta", "Histórico de introdução e uso de chupeta na rotina do lactente", "Nunca usou, Recusou o uso, Já usou mas não usa mais, Usa chupeta, Não sabe", "Mantido o formato original do questionário."],
    ["idade_mae_cat", "faixa_etaria_mae", "Faixa etária da mãe baseada em fatores de risco clínico de domínio. Adolescente (10-19 anos), Jovem Adulta (20-29 anos), Adulta (30-34 anos), Maternidade Tardia (35+ anos)", "Adolescente, Jovem Adulta, Adulta, Maternidade Tardia", "Discretização orientada por especialista. Valores numéricos agrupados em faixas. Instâncias nulas originais (7 casos) foram descartadas."],
    ["inic_prenat", "inicio_prenatal", "Momento gestacional do início do acompanhamento pré-natal (Imputado por IA)", "<12 semanas (Precoce), ≥12 semanas (Tardio)", "Valores nulos preenchidos utilizando Algoritmo de Machine Learning (Árvore de Decisão). Atributos utilizados para predição: 'faixa_renda_familiar', 'regiao_residencia', 'faixa_etaria_mae'."],
    ["", "inicio_precoce_amamentacao", "Identifica se o recém-nascido foi amamentado na primeira hora de vida", "Sim, Não", "Atributo derivado (Feature Engineering). Gerado a partir do tempo declarado até a primeira mamada (k13_tempo_medida e k12_tempo)."],
    ["j04_vive", "reside_com_parceiro", "Indica se a mãe coabita com parceiro, cônjuge ou companheiro fixo", "Sim, Não", "Mantido o formato original do questionário."],
    ["j06_ocupacao", "situacao_laboral_mae", "Situação laboral, contratual e de inserção no mercado de trabalho da mãe", "Trabalho regular, Trabalho irregular (bicos), Desempregado, Fora do mercado", "Mantido o formato original do questionário."],
    ["k03_prenatal", "realizou_prenatal", "Identifica se a gestante realizou consultas de pré-natal", "Sim, Não, Não sabe/Não quis responder", "Mantido o formato original do questionário."],
    ["k15_recebeu", "recebeu_outro_leite", "Durante sua permanência na maternidade a criança recebeu algum outro leite que não fosse o leite de peito", "Sim, Não, Não sabe/Não quis responder", "Mantido o formato original do questionário."],
    ["k16_liquido", "oferta_outros_liquidos", "Nos primeiros dias após o parto, e antes que o seu leite estivesse descendo, foi dado algum outro líquido para a criança beber que não fosse leite materno", "Sim, Não, Desconhecido", "Tratamento de ruído textuais. Respostas variadas como 'Não sabe' foram agrupadas como 'Desconhecido'."],
    ["k241_utilizou_concha", "usou_concha_amamentacao", "Uso de acessório: concha de amamentação", "Sim, Não", "Mantido o formato binário original."],
    ["k242_utilizou_protetor", "usou_protetor_mamilo", "Uso de acessório: protetor de mamilo / intermediário de silicone", "Sim, Não", "Mantido o formato binário original."],
    ["k243_utilizou_bico", "usou_bico_artificial", "Uso de acessório: outros bicos intermediários artificiais", "Sim, Não", "Mantido o formato binário original."],
    ["k244_utilizou_bomba", "usou_bomba_extracao", "Uso de acessório: bomba extratora/ordenha de leite materno", "Sim, Não", "Mantido o formato binário original."],
    ["k245_utilizou_mamadeira", "usou_mamadeira", "Uso de utensílio: mamadeira para oferta de complementos", "Sim, Não", "Mantido o formato binário original."],
    ["k246_utilizou_sondinha", "usou_sondinha_relactacao", "Uso de técnica clínica: relactação ou sondinha no peito", "Sim, Não", "Mantido o formato binário original."],
    ["k247_utilizou_copo", "usou_copinho", "Uso de utensílio protetivo: copinho para alimentação líquida", "Sim, Não", "Mantido o formato binário original."],
    ["k248_utilizou_nao", "nao_usou_acessorios", "Sinaliza que a mãe declarou NÃO ter utilizado nenhum acessório da série K24", "Sim, Não", "Mantido o formato binário original."],
    ["k249_utilizou_nao_sabe", "uso_acessorios_ignorado", "Não sabe ou recusou-se a responder sobre o uso de bicos e acessórios", "Sim, Não", "Mantido o formato binário original."],
    ["k28_aleitamento", "busca_informacao_aleitamento", "Você acessa ou acessou a internet para buscar informações sobre aleitamento materno", "Não, Pouco, Mais ou menos, Muito, Não sabe", "Mantido o formato original do questionário."],
    ["num_consultas", "faixa_consultas_prenatal", "Volume total quantitativo de consultas pré-natais realizadas (Imputado por IA)", "1-3 consultas, 4-6 consultas, 7e+ consultas", "Valores nulos preenchidos utilizando Algoritmo de Machine Learning (Árvore de Decisão). Atributos utilizados para predição: 'faixa_renda_familiar', 'regiao_residencia', 'inicio_prenatal'."],
    ["peso_cat", "classificacao_peso_nascimento", "Classificação ponderal epidemiológica do recém-nascido ao nascer", "Extremo baixo peso, Muito baixo peso, Baixo peso, Peso adequado, Macrossomia", "Mantido o formato original do questionário."],
    ["peso_ig", "adequacao_peso_idade_gestacional", "Relação de adequação do peso de nascimento com a idade gestacional", "PIG (Pequeno), AIG (Adequado), GIG (Grande)", "Mantido o formato original do questionário."],
    ["q01_recebe_beneficio", "recebe_auxilio_governamental", "Indica se a unidade familiar recebe auxílios e benefícios sociais do governo", "Sim, Não", "Mantido o formato original do questionário."],
    ["q07_renda_faixa", "faixa_renda_familiar", "Faixa de rendimento familiar mensal líquido (Classes altas aglutinadas)", "Sem renda, Até R$ 1.000,00, De R$ 1.001 a R$ 5.000, R$ 5.001,00 ou mais", "Categorias de alta renda aglutinadas para melhorar a distribuição estatística da base."],
    ["", "raca_cor_mae", "Etnia / Raça autorreferida da mãe de forma padronizada", "Branca, Parda, Preta, Amarela, Indígena", "Limpeza e padronização das classes minoritárias."],
    ["tempo_1a_cat", "tempo_ate_primeira_mamada", "Intervalo de tempo decorrido do nascimento até a primeira mamada efetiva", "<=1 hora, >1 a <=24 horas, >24 horas", "Discretizado em faixas de tempo a partir de variável numérica."],
    ["vd_ebia_categ", "nivel_inseguranca_alimentar", "Classificação da Escala Brasileira de Insegurança Alimentar (EBIA) no domicílio ", "Segurança, Insegurança leve, Insegurança moderada, Insegurança grave", "Mantido o formato original do questionário."]
]

for row in dic_data:
    ws_dic.append(row)

# ATENÇÃO AQUI: max_col alterado para 5 para cobrir a nova coluna
for row_idx, row in enumerate(ws_dic.iter_rows(min_row=1, max_row=len(dic_data), min_col=1, max_col=5), start=1):
    for col_idx, cell in enumerate(row, start=1):
        cell.border = thin_border
        if row_idx == 1:
            cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            if col_idx == 2:
                cell.font = Font(name="Consolas", size=10.5, bold=True, color="2C3E50")
            else:
                cell.font = font_regular
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color=color_zebra, end_color=color_zebra, fill_type="solid")

# Ajuste de tamanho das colunas, incluindo a nova coluna 'E'
ws_dic.row_dimensions[1].height = 25
ws_dic.column_dimensions['A'].width = 28
ws_dic.column_dimensions['B'].width = 28
ws_dic.column_dimensions['C'].width = 60
ws_dic.column_dimensions['D'].width = 55  
ws_dic.column_dimensions['E'].width = 60  

# =========================================================================
# TAB 3: DADOS PROCESSADOS (A Base de Dados)
# =========================================================================
ws_data = wb.create_sheet(title="Base de Dados Processada")
ws_data.views.sheetView[0].showGridLines = True

# Escrever cabeçalhos
headers_data = list(df.columns)
ws_data.append(headers_data)

# Inserir as linhas de dados (Maneira otimizada linha a linha)
for row in df.itertuples(index=False):
    ws_data.append(list(row))

# Aplicar estilos na aba de dados
for col_idx in range(1, len(headers_data) + 1):
    cell = ws_data.cell(row=1, column=col_idx)
    cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Estilo das linhas de dados (Zebra e alinhamento)
for r_idx in range(2, len(df) + 2):
    fill_to_apply = PatternFill(start_color=color_zebra, end_color=color_zebra, fill_type="solid") if r_idx % 2 == 0 else None
    for c_idx in range(1, len(headers_data) + 1):
        cell = ws_data.cell(row=r_idx, column=c_idx)
        cell.font = font_regular
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if fill_to_apply:
            cell.fill = fill_to_apply

ws_data.row_dimensions[1].height = 26
ws_data.freeze_panes = "A2"
ws_data.auto_filter.ref = f"A1:{get_column_letter(len(headers_data))}{len(df)+1}"

for c_idx in range(1, len(headers_data) + 1):
    col_letter = get_column_letter(c_idx)
    ws_data.column_dimensions[col_letter].width = 24

# Salvar o arquivo final estruturado
nome_saida_excel = r'C:\Users\vitoria-vaz\estudos\UFU\projeto-graduacao\pg-amamentacao\dataset\dataset_dicionario_amamentacao_pronto.xlsx'
wb.save(nome_saida_excel)
print(f"✅ Sucesso: O arquivo {nome_saida_excel} gerado com as observações de metadados!")